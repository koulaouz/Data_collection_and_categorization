import store1
import store2
import store3
import openpyxl
import pandas as pd
import re
import openpyxl
import multiprocessing
import subprocess

def runScript(s,i):
    s(i)

indexes = [0,10000,20000] # each store has drastically less than 10k products

if __name__ == "__main__":
    scripts = [store1.Main, store2.Main, store3.Main]

    processes = []

    for script, index in zip(scripts,indexes):
        process = multiprocessing.Process(target=runScript, args=(script,indexes))
        processes.append(process)
        process.start()

    for process in processes:
        process.join()


# Now the info is collected and we can process the data
workbook = openpyxl.load_workbook("total_2.xlsx")
worksheet = workbook.active
worksheet.delete_rows(2, worksheet.max_row)
workbook.save("total_2.xlsx")

df1 = pd.read_excel("store1.xlsx")
df2 = pd.read_excel("store2.xlsx")
df3 = pd.read_excel("store3.xlsx")

# Merge under a single file
merged_df = pd.concat([df1,df2,df3], ignore_index=False)
merged_df.to_excel('total_2.xlsx', index=False)

# Price normalization
workbook2 = openpyxl.load_workbook("total_2.xlsx")
worksheet2 = workbook2.active
for i in range(2, worksheet2.max_row+1):
    cell_obj = worksheet2.cell(row=i, column=4)
    priceValue = cell_obj.value
    match = re.search(r"\d+\,\d+|\d+\.\d+", priceValue)
    if match:
        priceValue = match.group()
        priceValue = priceValue.replace(",",".")
        cell_obj.value = priceValue

workbook2.save('total_2.xlsx')

# Data is collected and normalized. Continue with predicting categories.
subprocess.run(["python","predict_categories.py"])